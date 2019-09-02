# подготовительные шаги:
# 1. скачать и установить последнюю версию https://www.python.org/downloads/ . при установке убедиться, что проставлена галочка "добавить путь в PATH"
# 2. установить openpyxl через коммандную строку windows: pip install openpyxl
# 3. установить visual studio code 
# 4. установить расширения для visual studio code: python, code runner
# 5. убедиться, что расширение python использует системный интерпретатор, установленный на шаге 1 (view - command pallete - pyhton: select interpreted - ~applData)
# 6. в xlsx_input прописать полный путь до xlsx файла - !!! через /, а не \, например: c:/tmp/1.xlsx
# 7. запустить скрипт - ПКМ на редакторе - Run code

from openpyxl import load_workbook

topleft_row = 2
topleft_column = 1

xlsx_input = load_workbook(filename = "C:/git/csv2edt_converter/layers/sample.xlsx")
file_out = open("result.txt", "w")

xlsx_sheet = xlsx_input.active

curr_row = topleft_row
while True:   
    str_full_line = ""

    cell_num = xlsx_sheet.cell(row=curr_row, column=topleft_column)
    if cell_num.value == None:
        print("\nfound end of data on line {}".format(curr_row))
        break

    str_num = '{:>5}'.format(cell_num.value)
    str_full_line += str_num

    cell_deg = xlsx_sheet.cell(row=curr_row, column=topleft_column+1)
    str_deg = '{:>5}'.format(cell_deg.value)
    str_full_line += str_deg

    cell_H = xlsx_sheet.cell(row=curr_row, column=topleft_column+2)
    str_H = '{:>6}'.format(cell_H.value)
    str_full_line += str_H

    cell_start = xlsx_sheet.cell(row=curr_row, column=topleft_column+3)
    str_start = '{:>22}'.format(cell_start.value)
    str_full_line += str_start

    cell_Ro = xlsx_sheet.cell(row=curr_row, column=topleft_column+4)
    str_Ro = '{:11.1f}'.format(cell_Ro.value)
    str_full_line += str_Ro

    cell_Vs = xlsx_sheet.cell(row=curr_row, column=topleft_column+5)
    str_Vs = '{:>16}'.format(cell_Vs.value)
    str_full_line += str_Vs

    str_full_line = str_full_line.replace(" 0.", "  .")
    str_full_line = str_full_line.replace(".0 ", "   ")

    file_out.write(str_full_line)
    file_out.write('\n')

    curr_row += 1
    print(str_full_line)
